import streamlit as st
import pandas as pd
import requests
import json
from pandas import ExcelWriter
from collections import OrderedDict
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# ---------------- SUPABASE CONFIG ----------------
SUPABASE_URL = st.secrets["SUPABASE"]["URL"]
SUPABASE_KEY = st.secrets["SUPABASE"]["KEY"]
TABLE_NAME = "Volleyball_events"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json"
}

# ---------------- STREAMLIT CONFIG ----------------
st.set_page_config(page_title="🏐 Volleyball Event Dashboard", layout="wide")

# Initialize all session_state variables
for key in [
    "video_url", "game_name", "set_number",
    "selected_player", "selected_event",
    "selected_outcome", "attack_type", "set_to"
]:
    st.session_state.setdefault(key, "")

# ---------------- VIDEO INPUT ----------------
st.session_state.video_url = st.text_input(
    "🎥 YouTube Video URL",
    value=st.session_state.video_url,
    placeholder="https://www.youtube.com/watch?v=example",
    key="video_url_input"
)
if st.session_state.video_url:
    st.video(st.session_state.video_url)

# ---------------- GAME INFO ----------------
st.session_state.game_name = st.text_input(
    "🏆 Enter Game Name",
    value=st.session_state.game_name,
    placeholder="e.g. Blich vs Ramat Gan",
    key="game_name_input"
)

st.session_state.set_number = st.selectbox(
    "🎯 Select Set Number (optional)",
    ["", "1st Set", "2nd Set", "3rd Set", "4th Set", "5th Set"],
    index=["", "1st Set", "2nd Set", "3rd Set", "4th Set", "5th Set"].index(st.session_state.set_number)
    if st.session_state.set_number in ["", "1st Set", "2nd Set", "3rd Set", "4th Set", "5th Set"] else 0,
    key="set_number_select"
)

# ---------------- HELPER ----------------
def horizontal_radio(label, options, session_key):
    """Creates a radio button that persists in session_state."""
    current_value = st.session_state.get(session_key, options[0])
    value = st.radio(
        label,
        options,
        index=options.index(current_value) if current_value in options else 0,
        horizontal=True,
        key=session_key  # Syncs automatically to session_state
    )
    return value

def extract_category(event_value):
    if "(" in event_value:
        return event_value.split("(")[0].strip()
    return event_value

def auto_adjust_columns(writer, sheet_name):
    """Adjust column widths automatically."""
    ws = writer.sheets[sheet_name]
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


def export_player_excel(df, player_name):
    player_df = df[df["player"] == player_name].copy()
    if player_df.empty:
        st.error("No data for this player.")
        return

    player_df["category"] = player_df["event"].apply(extract_category)
    output_path = f"/tmp/{player_name}_volleyball_report.xlsx"

    # Outcome order per category (best → worst)
    OUTCOME_ORDER = {
        "Defense": ["Perfect", "Good", "Neutral", "Bad", "Overpass", "Failure"],
        "Dig": ["Perfect", "Good", "Neutral", "Bad"],
        "Receive": ["Perfect", "Good", "Neutral", "Bad", "Aced"]
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overall_summary = []

        for category in sorted(player_df["category"].unique()):
            cat_df = player_df[player_df["category"] == category]
            sheet_name = category[:31]

            # -------- Outcome statistics --------
            outcome_stats = cat_df.groupby("outcome").size().reset_index(name="count")
            total = outcome_stats["count"].sum()
            outcome_stats["percentage"] = (outcome_stats["count"] / total * 100).round(1)
            outcome_stats.loc[len(outcome_stats)] = ["TOTAL", total, 100.0]

            # Custom order
            if category in OUTCOME_ORDER:
                outcome_stats["order"] = outcome_stats["outcome"].apply(
                    lambda x: OUTCOME_ORDER[category].index(x) if x in OUTCOME_ORDER[category] else 999
                )
                outcome_stats = outcome_stats.sort_values("order").drop(columns="order")
            else:
                outcome_stats = outcome_stats.sort_values("count", ascending=False)

            outcome_stats.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

            # -------- Per-game table --------
            startrow_game = len(outcome_stats) + 3
            if "game_name" in cat_df.columns:
                game_stats = cat_df.groupby(["game_name", "outcome"]).size().reset_index(name="count")
                pivot_game = game_stats.pivot(index="game_name", columns="outcome", values="count").fillna(0)

                # Order columns in per-game table according to OUTCOME_ORDER
                if category in OUTCOME_ORDER:
                    ordered_cols = [c for c in OUTCOME_ORDER[category] if c in pivot_game.columns]
                    # Append any other unexpected columns at the end
                    ordered_cols += [c for c in pivot_game.columns if c not in ordered_cols]
                    pivot_game = pivot_game[ordered_cols]

                pivot_game.to_excel(writer, sheet_name=sheet_name, startrow=startrow_game)

            # -------- Per-category graph --------
            ws = writer.book[sheet_name]
            outcome_rows = len(outcome_stats) + 1
            per_game_rows = len(cat_df["game_name"].unique()) + 3 if "game_name" in cat_df.columns else 0
            startrow_chart = outcome_rows + per_game_rows + 5

            pivot_percent = cat_df.groupby(["game_name", "outcome"]).size().reset_index(name="count")
            total_per_game = pivot_percent.groupby("game_name")["count"].sum().reset_index()
            pivot_percent = pivot_percent.merge(total_per_game, on="game_name", suffixes=("", "_total"))
            pivot_percent["percentage"] = (pivot_percent["count"] / pivot_percent["count_total"] * 100).round(1)
            pivot_percent = pivot_percent.pivot(index="game_name", columns="outcome", values="percentage").fillna(0)

            # Order columns for graph
            if category in OUTCOME_ORDER:
                ordered_cols = [c for c in OUTCOME_ORDER[category] if c in pivot_percent.columns]
                ordered_cols += [c for c in pivot_percent.columns if c not in ordered_cols]
                pivot_percent = pivot_percent[ordered_cols]

            x_labels = list(pivot_percent.index)
            x = range(len(x_labels))
            colors = plt.cm.tab10.colors

            fig, ax = plt.subplots(figsize=(10, 4))
            for i, outcome_col in enumerate(pivot_percent.columns):
                y = pivot_percent[outcome_col].values
                ax.plot(x, y, marker='o', label=outcome_col, color=colors[i % len(colors)])
                for xi, yi in zip(x, y):
                    ax.text(xi, yi + 0.5, f"{yi}%", ha='center', va='bottom', fontsize=8)

            ax.set_title(f"{category} Performance (%)", fontsize=12)
            ax.set_ylabel("Percentage (%)")
            ax.set_ylim(0, 100)
            ax.grid(True, linestyle='--', alpha=0.5)
            ax.legend(title="Outcome", fontsize=8)
            ax.set_xticks(range(len(x_labels)))
            ax.set_xticklabels([f"Game {i+1}" for i in range(len(x_labels))], rotation=0)
            plt.tight_layout()

            img_data = BytesIO()
            plt.savefig(img_data, format="png", dpi=150)
            plt.close(fig)
            img_data.seek(0)
            img = XLImage(img_data)
            img.anchor = f"A{startrow_chart}"
            ws.add_image(img)

            # Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

            # Collect overall summary for this category
            for _, row in outcome_stats.iterrows():
                if row["outcome"] != "TOTAL":
                    overall_summary.append({
                        "Category": category,
                        "Outcome": row["outcome"],
                        "Count": row["count"],
                        "Percentage": row["percentage"]
                    })

        # -------- Summary Sheet --------
        summary_ws_row = 0
        wb = writer.book
        summary_sheet = wb.create_sheet("Summary")

        for category in sorted(player_df["category"].unique()):
            cat_summary = [s for s in overall_summary if s["Category"] == category]
            if not cat_summary:
                continue
            summary_df = pd.DataFrame(cat_summary)
            # Add a total row per category
            total_count = summary_df["Count"].sum()
            total_percentage = 100.0
            summary_df.loc[len(summary_df)] = ["TOTAL", "", total_count, total_percentage]

            summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=summary_ws_row)
            summary_ws_row += len(summary_df) + 3  # space between tables

    st.success("✅ Excel report created!")
    with open(output_path, "rb") as f:
        st.download_button(
            "⬇️ Download Excel",
            f,
            file_name=f"{player_name}_volleyball_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ---------------- PLAYER SELECTION ----------------
player = horizontal_radio("### 🏐 Select Player", 
    ["", "Ori", "Ofir", "Beni", "Hillel", "Shak", "Omer Saar", "Omer", "Karat", "Lior", "Yonatan", "Ido", "Royi"], 
    "selected_player"
)

# ---------------- EVENT SELECTION ----------------
event = horizontal_radio("### ⚡ Select Event", 
    ["", "Serve", "Attack", "Block", "Receive", "Dig", "Set", "Defense"], 
    "selected_event"
)

# ---------------- SUBCHOICES ----------------
attack_type = None
set_to = None

event_outcomes = {
    "Serve": ["Ace", "Out", "Net", "In", "Off System", "Overpass"],
    "Attack": ["Blockout", "Out", "Net", "In Play", "Off System", "Success"],
    "Block": ["Blockout", "Touch", "Kill", "Softblock", "Error"],
    "Receive": ["Perfect", "Good", "Neutral", "Bad","Aced"],
    "Dig": ["Good", "Neutral", "Bad"],
    "Set": ["0 Blockers", "1 Blocker", "2 Blocker", "Overset"],
    "Defense": ["Good", "Neutral", "Bad", "Overpass", "Failure"]
}

if event == "Attack":
    attack_type = horizontal_radio("### ⚡ Attack Type", ["", "Free Ball", "Tip", "Hole", "Spike"], "attack_type")
elif event == "Set":
    set_to = horizontal_radio("### 🧱 Set To", ["", "Position 1", "Position 2", "Position 3", "Position 4", "Position 6"], "set_to")

# ---------------- OUTCOME ----------------
base_outcomes = event_outcomes.get(event, [])
if event == "Attack":
    if st.session_state.get("last_attack_type") != attack_type:
        st.session_state.selected_outcome = ""
    st.session_state.last_attack_type = attack_type


if event == "Attack" and attack_type == "Spike":
    outcome_options = base_outcomes + ["Hard Blocked", "Soft Blocked"]
else:
    outcome_options = base_outcomes

outcome = (
    horizontal_radio("### 🎯 Select Outcome", [""] + outcome_options, "selected_outcome")
    if outcome_options else None
)
free_text = st.text_area(
    "📝 Additional Notes (optional)",
    placeholder="Write anything you want to attach to this event...",
    key="free_text"
)
# ---------------- SUPABASE OPS ----------------
def save_event():
    extra_info = attack_type if attack_type else set_to
    data = {
        "player": player,
        "event_category": event,
        "attack_type": attack_type if event == "Attack" else None,
        "outcome": outcome,
        "video_url": st.session_state.video_url,
        "game_name": st.session_state.game_name if st.session_state.game_name else "No Game Entered",
        "set_number": st.session_state.set_number if st.session_state.set_number else None,
        "notes": free_text
    }

    response = requests.post(f"{SUPABASE_URL}/rest/v1/{TABLE_NAME}", headers=HEADERS, data=json.dumps(data))
    if response.status_code in [200, 201]:
        st.success(f"✅ Saved: {player} | {event} | {extra_info if extra_info else ''} | {outcome}")
    else:
        st.error(f"❌ Failed to save: {response.text}")

def load_events():
    response = requests.get(f"{SUPABASE_URL}/rest/v1/{TABLE_NAME}?select=*", headers=HEADERS)
    if response.status_code == 200:
        return pd.DataFrame(response.json())
    else:
        st.error(f"Failed to load events: {response.text}")
        return pd.DataFrame()

def update_event(row_id, updated_data):
    url = f"{SUPABASE_URL}/rest/v1/{TABLE_NAME}?id=eq.{row_id}"
    response = requests.patch(url, headers=HEADERS, data=json.dumps(updated_data))
    if response.status_code not in [200, 204]:
        st.error(f"❌ Failed to update row {row_id}: {response.text}")

def delete_event(row_id):
    url = f"{SUPABASE_URL}/rest/v1/{TABLE_NAME}?id=eq.{row_id}"
    response = requests.delete(url, headers=HEADERS)
    if response.status_code not in [200, 204]:
        st.error(f"❌ Failed to delete row {row_id}: {response.text}")

# ---------------- SAVE BUTTON ----------------
if st.button("💾 Save Event", use_container_width=True):
    if player and event and outcome:
        save_event()
    else:
        st.error("⚠️ Please select a player, event, and outcome before saving.")


# ---------------- LOGGED EVENTS ----------------
df = load_events()

if not df.empty:
    # Sort by timestamp or id descending to show newest first
    sort_col = "timestamp" if "timestamp" in df.columns else "id"
    df = df.sort_values(sort_col, ascending=False)

    with st.expander("🔍 Filter"):
        sel_game = st.multiselect("Game", df["game_name"].dropna().unique())
        sel_player = st.multiselect("Player", df["player"].unique())
        sel_event = st.multiselect("Event", df["event"].unique())

        # Apply filters
        if sel_game:
            df = df[df["game_name"].isin(sel_game)]
        if sel_player:
            df = df[df["player"].isin(sel_player)]
        if sel_event:
            df = df[df["event"].isin(sel_event)]

        # Re-sort after filtering to keep newest on top
        df = df.sort_values(sort_col, ascending=False)

    # -------- Editable + Deletable Table --------
    df_display = df.drop(columns=["timestamp"], errors="ignore").copy()
    df_display["Delete?"] = False  # Add a checkbox column

    edited_df = st.data_editor(
        df_display,
        num_rows="fixed",
        use_container_width=True,
        key="editor",
    )

    # Handle edits
    if st.button("💾 Save All Changes", use_container_width=True):
        for i, row in edited_df.iterrows():
            original = df.loc[df["id"] == row["id"]].iloc[0]
            changes = {col: row[col] for col in df.columns if col in row and row[col] != original[col]}
            if changes:
                update_event(row["id"], changes)
        st.success("✅ All edits saved!")
        st.rerun()

    # Handle deletes
    delete_ids = edited_df.loc[edited_df["Delete?"], "id"].tolist()
    if delete_ids and st.button("🗑️ Delete Selected Rows", use_container_width=True):
        for row_id in delete_ids:
            delete_event(row_id)
        st.rerun()

    # CSV Export
    st.download_button(
        "⬇️ Download CSV",
        df.to_csv(index=False).encode("utf-8"),
        "volleyball_events.csv",
        "text/csv"
    )

    st.divider()
    st.subheader("📊 Player Statistics Export")
    
    player_for_export = st.selectbox(
        "Select player to export",
        sorted(df["player"].dropna().unique())
    )
    
    if st.button("⬇️ Download Player Excel Report"):
        export_player_excel(df, player_for_export)

else:
    st.info("No events logged yet.")

